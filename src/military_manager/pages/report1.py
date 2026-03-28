"""Report 1 — דו"ח 1 — Company-level presence report.

Similar to the Excel "סיכום פלוגתי" sheet:
  - Rows: soldiers grouped by sub-unit (מחלקה)
  - Columns: dates
  - Cells: daily status (בבסיס, חופש, התייצב, etc.)
  - Bottom summary: counts by role category (officers, drivers, fighters, etc.)
  - Color-coded statuses
  - Multi-day forward view
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from collections import defaultdict

import pandas as pd
import streamlit as st

from military_manager.components.navigation import render_page_header
from military_manager.components.filters import period_guard, sub_unit_filter
from military_manager.config import IRRELEVANT_UNIT
from military_manager.services.soldier_service import get_period_soldiers, get_sub_units
from military_manager.services.status_service import (
    get_daily_status_grid,
    get_daily_counts,
    set_status,
    set_status_notes,
    bulk_set_status,
    bulk_clear_status,
    get_soldier_status_history,
)
from military_manager.services.period_service import get_status_options
from military_manager.database import get_session, Request

# ── Status categories for color coding ──
PRESENT_STATUSES = {"בבסיס", "התייצב", "חוזר מחופש", "סיפוח מאוחר", "צפוי להתייצב"}
AWAY_STATUSES = {"חופש", "יוצא לחופש", "יוצא לפיצול", "פיצול", "גימלים", "משתחרר"}
ABSENT_STATUSES = {"לא בשמפ", "נפקד"}

STATUS_COLORS = {
    "בבסיס": "#C8E6C9",       # light green
    "התייצב": "#A5D6A7",      # green
    "חוזר מחופש": "#E8F5E9",  # pale green
    "צפוי להתייצב": "#DCEDC8",
    "סיפוח מאוחר": "#F0F4C3",
    "חופש": "#FFCDD2",        # light red
    "יוצא לחופש": "#FFAB91",  # orange-red
    "פיצול": "#FFE0B2",       # light orange
    "יוצא לפיצול": "#FFE0B2",
    "גימלים": "#F8BBD0",      # pink
    "משתחרר": "#E1BEE7",      # purple
    "לא בשמפ": "#CFD8DC",     # grey
    "נפקד": "#FF8A80",        # red
}

# ── Role categories for the bottom summary ──
# EXACT: role must match exactly (after quote normalization)
# PREFIX: role must start with the pattern
from collections import OrderedDict

_ROLE_EXACT: dict[str, list[str]] = {
    "מ\"מים": ["מ\"מ"],
}

_ROLE_PREFIX = OrderedDict([
    ("מפקד גזרה", ["מ\"פ", "סמ\"פ"]),
    ("קצינים", ["רס\"פ", "ע.מ\"פ", "מהנדס", "קצין"]),
    ("סמלי מחלקה", ["סמל מחלקה"]),
    ("מ\"כים", ["מ\"כ"]),
    ("נהגים", ["נהג"]),
    ("לוחמים", ["לוחם", "מחלץ"]),
    ("חמ\"ל", ["חמל", "חמליסט", "קשר"]),
    ("חובשים", ["חובש"]),
])

# Combined ordered list for display
ROLE_CATEGORIES_ORDER = [
    "מפקד גזרה", "קצינים", "מ\"מים", "סמלי מחלקה", "מ\"כים",
    "נהגים", "לוחמים", "חמ\"ל", "חובשים",
]


def _normalize_role(val: str) -> str:
    return val.replace('"', '').replace("'", "").lower().strip()


def _get_role_category(soldier: dict) -> str:
    """Classify a soldier into a role category.
    Uses exact match first, then prefix match on the *role* field only."""
    role = (soldier.get("role", "") or "").strip()
    role_norm = _normalize_role(role)
    if not role_norm:
        return "אחר"

    # 1) Exact match takes priority
    for cat, patterns in _ROLE_EXACT.items():
        for p in patterns:
            if role_norm == _normalize_role(p):
                return cat

    # 2) Prefix match (role starts with pattern, optionally followed by space)
    for cat, patterns in _ROLE_PREFIX.items():
        for p in patterns:
            p_norm = _normalize_role(p)
            if role_norm == p_norm or role_norm.startswith(p_norm + " "):
                return cat

    return "אחר"

HEB_DAYS = {
    0: "ב", 1: "ג", 2: "ד", 3: "ה",
    4: "ו", 5: "ש", 6: "א",
}


def render():
    render_page_header("📊 דו\"ח 1", "סיכום פלוגתי — נוכחות חיילים לפי תאריכים")

    period = period_guard()
    if not period:
        return

    pid = period["id"]

    try:
        p_start = datetime.strptime(period["start_date"], "%Y-%m-%d").date()
        p_end = datetime.strptime(period["end_date"], "%Y-%m-%d").date()
    except (ValueError, KeyError):
        p_start = date.today()
        p_end = date.today() + timedelta(days=21)

    # ── Missing report alert — tomorrow not filled by 15:00 ──
    _render_tomorrow_alert(pid, p_start, p_end)

    tab_report, tab_edit, tab_summary, tab_groups, tab_students = st.tabs([
        "📊 דו\"ח נוכחות",
        "✏️ עדכון סטטוס מהיר",
        "📈 סיכום כוח אדם",
        "📊 אחוזים לפי קבוצות",
        "🎓 סטודנטים",
    ])

    with tab_report:
        _render_presence_report(pid, p_start, p_end)

    with tab_edit:
        _render_quick_status_edit(pid, p_start, p_end)

    with tab_summary:
        _render_manpower_summary(pid, p_start, p_end)

    with tab_groups:
        _render_group_percentage_report(pid, p_start, p_end)

    with tab_students:
        _render_student_report(pid, p_start, p_end)


def _render_tomorrow_alert(pid: int, p_start, p_end):
    """Show alert if tomorrow's report is not yet filled and it's past 15:00."""
    now = datetime.now()
    tomorrow = date.today() + timedelta(days=1)

    # Only alert if tomorrow is within the period
    if tomorrow < p_start or tomorrow > p_end:
        return

    # Only show alert after 15:00
    if now.hour < 15:
        return

    # Check how many soldiers have a status set for tomorrow
    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)
    if not soldiers:
        return

    grid = get_daily_status_grid(pid, tomorrow, tomorrow)
    status_map = grid.get("statuses", {}) if grid else {}

    filled = 0
    total = len(soldiers)
    missing_by_unit: dict[str, list[str]] = defaultdict(list)

    for s in soldiers:
        sid = s["soldier_id"]
        dk = tomorrow.isoformat()
        key = f"{sid}_{dk}"
        status = status_map.get(key, "")
        if status:
            filled += 1
        else:
            unit = s.get("sub_unit", "ללא מחלקה")
            missing_by_unit[unit].append(s.get("full_name", "?"))

    missing = total - filled

    if missing == 0:
        return  # All filled — no alert

    pct_filled = round(filled / total * 100) if total > 0 else 0

    # Build detailed alert
    unit_details = []
    for u, names in sorted(missing_by_unit.items()):
        unit_details.append(f"**{u}** ({len(names)} חסרים): {', '.join(names[:5])}{'...' if len(names) > 5 else ''}")

    alert_md = (
        f"⚠️ **דו\"ח 1 ליום מחר ({tomorrow.strftime('%d/%m')}) עדיין לא מלא!**\n\n"
        f"מולאו **{filled}/{total}** חיילים ({pct_filled}%) — "
        f"**חסרים {missing} חיילים**\n\n"
        + "\n\n".join(unit_details)
        + "\n\n**נדרש למלא את הדוח בהקדם.**"
    )
    st.warning(alert_md)


# ╔══════════════════════════════════════════════════════════════╗
# ║                   PRESENCE REPORT TABLE                      ║
# ╚══════════════════════════════════════════════════════════════╝

def _render_presence_report(pid: int, p_start, p_end):
    """Main report — color-coded HTML table like Excel סיכום פלוגתי."""
    st.markdown("### 📊 דו\"ח נוכחות — סיכום פלוגתי")

    # Date range selector
    col1, col2, col3 = st.columns([2, 2, 2])
    with col1:
        view_start = st.date_input(
            "מתאריך",
            value=min(max(p_start, date.today()), p_end),
            min_value=p_start,
            max_value=p_end,
            key="r1_start",
        )
    with col2:
        default_end = min(view_start + timedelta(days=13), p_end)
        view_end = st.date_input(
            "עד תאריך",
            value=default_end,
            min_value=view_start,
            max_value=p_end,
            key="r1_end",
        )
    with col3:
        unit_filter = sub_unit_filter(pid, key="r1_unit")

    num_days = (view_end - view_start).days + 1
    st.caption(f"📅 {num_days} ימים")

    # Build date list
    dates = []
    current = view_start
    while current <= view_end:
        dates.append(current)
        current += timedelta(days=1)

    # Get all soldiers
    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)
    if unit_filter:
        soldiers = [s for s in soldiers if s.get("sub_unit") == unit_filter]

    if not soldiers:
        st.info("אין חיילים להצגה")
        return

    # Get status grid
    grid = get_daily_status_grid(pid, view_start, view_end, sub_unit=unit_filter)
    status_map = grid.get("statuses", {}) if grid else {}
    notes_map = grid.get("notes", {}) if grid else {}

    # Group soldiers by sub_unit
    units = defaultdict(list)
    for s in soldiers:
        units[s.get("sub_unit", "ללא מחלקה")].append(s)

    # Sort units
    sorted_units = sorted(units.keys())

    # ── Presence counter banner ──
    _render_presence_banner(soldiers, dates, status_map)

    # ── Export to Excel button ──
    st.markdown("---")
    excel_bytes = _build_report_excel(sorted_units, units, dates, status_map, soldiers)
    st.download_button(
        label="📥 ייצוא לאקסל",
        data=excel_bytes,
        file_name=f"דוח_1_{view_start.strftime('%d_%m_%Y')}_{view_end.strftime('%d_%m_%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="r1_export_excel",
    )

    # Build HTML table
    html = _build_report_html(sorted_units, units, dates, status_map, notes_map)
    st.markdown(html, unsafe_allow_html=True)

    # Notes summary
    if notes_map:
        st.markdown("---")
        _render_notes_summary(notes_map, soldiers, dates)

    # Legend
    st.markdown("---")
    _render_legend()

    # Bottom summary — counts by role category per day
    st.markdown("---")
    _render_role_summary_table(soldiers, dates, status_map)


def _build_report_html(
    sorted_units: list,
    units: dict,
    dates: list,
    status_map: dict,
    notes_map: dict | None = None,
) -> str:
    """Build the full HTML report table."""
    html = """
    <style>
    .r1-table { width:100%; border-collapse:separate; border-spacing:0;
        direction:rtl; font-size:11px; }
    .r1-table th { background:#1B5E20; color:white; padding:4px 3px; text-align:center;
        border:1px solid #aaa; position:sticky; top:0; z-index:3; font-size:10px; }
    .r1-table td { padding:3px 2px; border:1px solid #ddd; text-align:center;
        min-width:55px; font-size:10px; white-space:nowrap; }
    .r1-name { text-align:right !important; padding-right:6px !important;
        white-space:nowrap; min-width:100px; font-weight:500;
        position:sticky; right:0; z-index:2; background:#fff; }
    .r1-role { text-align:center !important; font-size:9px; color:#555;
        min-width:50px; position:sticky; right:100px; z-index:2; background:#fff; }
    /* Attached soldier (מסופח) — light peach tint */
    .r1-name-attached { text-align:right !important; padding-right:6px !important;
        white-space:nowrap; min-width:100px; font-weight:500;
        position:sticky; right:0; z-index:2; background:#FFF3E0; }
    .r1-role-attached { text-align:center !important; font-size:9px; color:#555;
        min-width:50px; position:sticky; right:100px; z-index:2; background:#FFF3E0; }
    .r1-unit-header { background:#E8F5E9 !important; font-weight:bold;
        text-align:right !important; padding-right:8px !important;
        font-size:12px; color:#1B5E20; position:sticky; right:0; z-index:2; }
    .r1-summary { background:#FFF9C4; font-weight:bold; font-size:11px; }
    .r1-total { background:#BBDEFB; font-weight:bold; font-size:11px; }
    /* Sticky header columns (name+role) for headers too */
    .r1-table th:first-child { position:sticky; right:0; z-index:4; background:#1B5E20; }
    .r1-table th:nth-child(2) { position:sticky; right:100px; z-index:4; background:#1B5E20; }
    </style>
    <div style="overflow-x:auto; max-height:600px; overflow-y:auto; position:relative;">
    <table class="r1-table">
    """

    # Header row 1 — day names
    html += "<tr><th>שם</th><th>תפקיד</th>"
    for d in dates:
        day_heb = HEB_DAYS.get(d.weekday(), "")
        html += f"<th>{day_heb}</th>"
    html += "</tr>"

    # Header row 2 — dates
    html += "<tr><th></th><th></th>"
    for d in dates:
        html += f"<th>{d.strftime('%d/%m')}</th>"
    html += "</tr>"

    # Per-unit rows
    for unit_name in sorted_units:
        unit_soldiers = units[unit_name]

        # Unit header row
        colspan = len(dates) + 2
        html += f'<tr><td class="r1-unit-header" colspan="{colspan}">'
        html += f'📁 {unit_name} ({len(unit_soldiers)} חיילים)</td></tr>'

        for s in unit_soldiers:
            sid = s["soldier_id"]
            name = s.get("full_name", "")
            role = s.get("role", "") or ""
            attached = s.get("is_attached", False)

            # Attached soldiers get a peach-tinted row
            name_cls = "r1-name-attached" if attached else "r1-name"
            role_cls = "r1-role-attached" if attached else "r1-role"
            attached_suffix = " 📎" if attached else ""

            html += "<tr>"
            html += f'<td class="{name_cls}">{name}{attached_suffix}</td>'
            html += f'<td class="{role_cls}">{role}</td>'

            for d in dates:
                key = f"{sid}_{d.isoformat()}"
                status = status_map.get(key, "")
                note = (notes_map or {}).get(key, "")
                bg = STATUS_COLORS.get(status, "#fff")
                short = _short_status(status) if status else ""
                tooltip = status
                if note:
                    tooltip += f" | {note}"
                    short += " 📝"
                html += f'<td style="background:{bg};" title="{tooltip}">{short}</td>'

            html += "</tr>"

    html += "</table></div>"
    return html


# ── Excel color mapping (hex → openpyxl PatternFill) ──
_EXCEL_STATUS_FILLS: dict[str, str] = {
    "בבסיס": "C8E6C9",
    "התייצב": "A5D6A7",
    "חוזר מחופש": "E8F5E9",
    "צפוי להתייצב": "DCEDC8",
    "סיפוח מאוחר": "F0F4C3",
    "חופש": "FFCDD2",
    "יוצא לחופש": "FFAB91",
    "פיצול": "FFE0B2",
    "יוצא לפיצול": "FFE0B2",
    "גימלים": "F8BBD0",
    "משתחרר": "E1BEE7",
    "לא בשמפ": "CFD8DC",
    "נפקד": "FF8A80",
}


def _build_report_excel(
    sorted_units: list,
    units: dict,
    dates: list,
    status_map: dict,
    soldiers: list[dict],
) -> bytes:
    """Build a styled Excel workbook for the presence report and role summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Presence report ──
    ws = wb.active
    ws.title = "דוח נוכחות"
    ws.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    unit_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    unit_font = Font(bold=True, color="1B5E20", size=11)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    attached_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # Row 1: header — day names
    ws.cell(row=1, column=1, value="שם").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = border
    ws.cell(row=1, column=2, value="תפקיד").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).border = border
    for ci, d in enumerate(dates, start=3):
        c = ws.cell(row=1, column=ci, value=HEB_DAYS.get(d.weekday(), ""))
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Row 2: dates
    ws.cell(row=2, column=1, value="").fill = header_fill
    ws.cell(row=2, column=1).border = border
    ws.cell(row=2, column=2, value="").fill = header_fill
    ws.cell(row=2, column=2).border = border
    for ci, d in enumerate(dates, start=3):
        c = ws.cell(row=2, column=ci, value=d.strftime("%d/%m"))
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    row_idx = 3
    for unit_name in sorted_units:
        unit_soldiers = units[unit_name]
        # Unit header row — merged
        ws.merge_cells(
            start_row=row_idx, start_column=1,
            end_row=row_idx, end_column=len(dates) + 2,
        )
        c = ws.cell(row=row_idx, column=1, value=f"{unit_name} ({len(unit_soldiers)} חיילים)")
        c.font = unit_font
        c.fill = unit_fill
        c.alignment = right_align
        c.border = border
        row_idx += 1

        for s in unit_soldiers:
            sid = s["soldier_id"]
            name = s.get("full_name", "")
            role = s.get("role", "") or ""
            attached = s.get("is_attached", False)

            name_cell = ws.cell(row=row_idx, column=1, value=name + (" 📎" if attached else ""))
            name_cell.alignment = right_align
            name_cell.border = border
            name_cell.font = Font(bold=False, size=10)
            if attached:
                name_cell.fill = attached_fill

            role_cell = ws.cell(row=row_idx, column=2, value=role)
            role_cell.alignment = center
            role_cell.border = border
            role_cell.font = Font(size=9, color="555555")
            if attached:
                role_cell.fill = attached_fill

            for ci, d in enumerate(dates, start=3):
                key = f"{sid}_{d.isoformat()}"
                status = status_map.get(key, "")
                c = ws.cell(row=row_idx, column=ci, value=status)
                c.alignment = center
                c.border = border
                fill_hex = _EXCEL_STATUS_FILLS.get(status)
                if fill_hex:
                    c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

            row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for ci in range(3, len(dates) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── Sheet 2: Role summary ──
    ws2 = wb.create_sheet("סיכום כוח אדם")
    ws2.sheet_view.rightToLeft = True

    categories = ROLE_CATEGORIES_ORDER + ["אחר"]
    summary_data: dict[str, dict] = {cat: {} for cat in categories}
    total_by_date: dict = {}

    for d in dates:
        total_present = 0
        total_all = 0
        for cat in categories:
            summary_data[cat][d] = {"present": 0, "total": 0}
        for s in soldiers:
            sid = s["soldier_id"]
            cat = _get_role_category(s)
            key = f"{sid}_{d.isoformat()}"
            status = status_map.get(key, "")
            summary_data[cat][d]["total"] += 1
            total_all += 1
            if status in PRESENT_STATUSES:
                summary_data[cat][d]["present"] += 1
                total_present += 1
        total_by_date[d] = {"present": total_present, "total": total_all}

    sum_header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    sum_header_font = Font(bold=True, color="FFFFFF", size=10)
    cat_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    total_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Header row
    ws2.cell(row=1, column=1, value="קטגוריה").font = sum_header_font
    ws2.cell(row=1, column=1).fill = sum_header_fill
    ws2.cell(row=1, column=1).border = border
    for ci, d in enumerate(dates, start=2):
        day_heb = HEB_DAYS.get(d.weekday(), "")
        c = ws2.cell(row=1, column=ci, value=f"{day_heb} {d.strftime('%d/%m')}")
        c.font = sum_header_font
        c.fill = sum_header_fill
        c.alignment = center
        c.border = border

    r = 2
    for cat in categories:
        if not any(summary_data[cat][d]["total"] > 0 for d in dates):
            continue
        c = ws2.cell(row=r, column=1, value=cat)
        c.font = Font(bold=True, size=10)
        c.fill = cat_fill
        c.alignment = right_align
        c.border = border

        for ci, d in enumerate(dates, start=2):
            pres = summary_data[cat][d]["present"]
            total = summary_data[cat][d]["total"]
            c = ws2.cell(row=r, column=ci, value=f"{pres}/{total}" if total else "—")
            c.alignment = center
            c.border = border
            if total > 0:
                pct = round(pres / total * 100)
                c.fill = green_fill if pct >= 70 else yellow_fill if pct >= 40 else red_fill
        r += 1

    # Total row
    c = ws2.cell(row=r, column=1, value='סה"כ')
    c.font = Font(bold=True, size=10)
    c.fill = total_fill
    c.alignment = right_align
    c.border = border
    for ci, d in enumerate(dates, start=2):
        pres = total_by_date[d]["present"]
        total = total_by_date[d]["total"]
        pct_str = f" ({round(pres/total*100)}%)" if total else ""
        c = ws2.cell(row=r, column=ci, value=f"{pres}/{total}{pct_str}")
        c.font = Font(bold=True)
        c.fill = total_fill
        c.alignment = center
        c.border = border

    ws2.column_dimensions["A"].width = 18
    for ci in range(2, len(dates) + 2):
        ws2.column_dimensions[get_column_letter(ci)].width = 14

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _short_status(status: str) -> str:
    """Shorten status label for compact display."""
    mapping = {
        "בבסיס": "בסיס",
        "התייצב": "התייצב",
        "חוזר מחופש": "חוזר",
        "צפוי להתייצב": "צפוי",
        "סיפוח מאוחר": "סיפוח",
        "חופש": "חופש",
        "יוצא לחופש": "יוצא",
        "פיצול": "פיצול",
        "יוצא לפיצול": "פיצול",
        "גימלים": "גימלים",
        "משתחרר": "משתחרר",
        "לא בשמפ": "לא בשמ\"פ",
        "נפקד": "נפקד",
    }
    return mapping.get(status, status[:6] if len(status) > 6 else status)


def _render_presence_banner(soldiers: list[dict], dates: list, status_map: dict):
    """Show a prominent banner with presence count per date."""
    total = len(soldiers)
    if not total:
        return

    today = date.today()
    # If today is in the date range, show it prominently
    if today in dates:
        present = 0
        for s in soldiers:
            key = f"{s['soldier_id']}_{today.isoformat()}"
            status = status_map.get(key, "")
            if status in PRESENT_STATUSES:
                present += 1
        pct = round(present / total * 100) if total else 0
        color = "#4CAF50" if pct >= 70 else ("#FF9800" if pct >= 50 else "#F44336")
        st.markdown(
            f'<div style="background:{color}18; border:2px solid {color}; border-radius:12px; '
            f'padding:12px 20px; text-align:center; margin-bottom:10px;">'
            f'<span style="font-size:1.4em; font-weight:bold; color:{color};">'
            f'🏕️ בבסיס היום: {present} מתוך {total} ({pct}%)'
            f'</span></div>',
            unsafe_allow_html=True,
        )
    else:
        # Show for the first date in range
        d = dates[0]
        present = 0
        for s in soldiers:
            key = f"{s['soldier_id']}_{d.isoformat()}"
            status = status_map.get(key, "")
            if status in PRESENT_STATUSES:
                present += 1
        pct = round(present / total * 100) if total else 0
        color = "#4CAF50" if pct >= 70 else ("#FF9800" if pct >= 50 else "#F44336")
        st.markdown(
            f'<div style="background:{color}18; border:2px solid {color}; border-radius:12px; '
            f'padding:12px 20px; text-align:center; margin-bottom:10px;">'
            f'<span style="font-size:1.4em; font-weight:bold; color:{color};">'
            f'🏕️ בבסיס ב-{d.strftime("%d/%m")}: {present} מתוך {total} ({pct}%)'
            f'</span></div>',
            unsafe_allow_html=True,
        )


def _render_notes_summary(
    notes_map: dict[str, str],
    soldiers: list[dict],
    dates: list,
):
    """Summary table grouping soldiers by their note text."""
    # Build soldier ID → name lookup
    id_to_name: dict[int, str] = {}
    for s in soldiers:
        sid = s["soldier_id"]
        id_to_name[sid] = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()

    # Set of soldier IDs in current view
    soldier_ids = set(id_to_name.keys())

    # Group: note_text → set of (soldier_name, date_str)
    from collections import defaultdict as _dd
    note_groups: dict[str, dict[str, list[str]]] = _dd(lambda: _dd(list))
    #  note_text -> soldier_name -> [dates]

    for key, note_text in notes_map.items():
        if not note_text or not note_text.strip():
            continue
        parts = key.split("_", 1)
        if len(parts) != 2:
            continue
        try:
            sid = int(parts[0])
        except ValueError:
            continue
        if sid not in soldier_ids:
            continue
        date_str = parts[1]  # ISO date
        normalized = note_text.strip()
        name = id_to_name.get(sid, str(sid))
        note_groups[normalized][name].append(date_str)

    if not note_groups:
        return

    st.markdown("### 📝 סיכום הערות")

    # Sort by number of soldiers descending
    sorted_notes = sorted(
        note_groups.items(),
        key=lambda x: len(x[1]),
        reverse=True,
    )

    for note_text, soldiers_dict in sorted_notes:
        count = len(soldiers_dict)
        with st.expander(f"📌 {note_text} — {count} חיילים"):
            rows = []
            for name in sorted(soldiers_dict.keys()):
                date_list = soldiers_dict[name]
                # Format dates nicely
                formatted_dates = ", ".join(
                    sorted(d.replace("-", "/") for d in date_list)
                )
                rows.append({"שם": name, "תאריכים": formatted_dates})
            import pandas as _pd
            df = _pd.DataFrame(rows)
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "שם": st.column_config.TextColumn("שם", width="medium"),
                    "תאריכים": st.column_config.TextColumn("תאריכים", width="large"),
                },
            )


def _render_legend():
    """Show color legend for status types."""
    st.markdown("**מקרא צבעים:**")
    cols = st.columns(6)
    items = list(STATUS_COLORS.items())
    for i, (status, color) in enumerate(items):
        cols[i % 6].markdown(
            f'<span style="background:{color}; padding:2px 8px; '
            f'border-radius:4px; font-size:12px;">{status}</span>',
            unsafe_allow_html=True,
        )
    st.markdown(
        '<span style="background:#FFF3E0; padding:2px 8px; border-radius:4px; '
        'font-size:12px;">📎 חייל מסופח</span>',
        unsafe_allow_html=True,
    )


def _render_role_summary_table(
    soldiers: list[dict],
    dates: list,
    status_map: dict,
):
    """Bottom summary table — per role category per day."""
    st.markdown("### 📋 סיכום כוח אדם לפי תפקיד")

    # Build summary: for each category × date → count present
    categories = ROLE_CATEGORIES_ORDER + ["אחר"]
    summary_data = {cat: {} for cat in categories}
    total_by_date = {}

    for d in dates:
        total_present = 0
        total_all = 0
        for cat in categories:
            summary_data[cat][d] = {"present": 0, "total": 0}

        for s in soldiers:
            sid = s["soldier_id"]
            cat = _get_role_category(s)
            key = f"{sid}_{d.isoformat()}"
            status = status_map.get(key, "")

            summary_data[cat][d]["total"] += 1
            total_all += 1

            if status in PRESENT_STATUSES:
                summary_data[cat][d]["present"] += 1
                total_present += 1

        total_by_date[d] = {"present": total_present, "total": total_all}

    # Build HTML for summary
    html = """
    <style>
    .r1s-table { width:100%; border-collapse:separate; border-spacing:0;
        direction:rtl; font-size:12px; }
    .r1s-table th { background:#1565C0; color:white; padding:5px 4px;
        text-align:center; border:1px solid #aaa; font-size:11px;
        position:sticky; top:0; z-index:3; }
    .r1s-table th:first-child { position:sticky; right:0; z-index:4;
        background:#1565C0; }
    .r1s-table td { padding:4px; border:1px solid #ddd; text-align:center;
        font-size:11px; }
    .r1s-cat { text-align:right !important; padding-right:8px !important;
        font-weight:bold; background:#E3F2FD; min-width:80px;
        position:sticky; right:0; z-index:2; }
    .r1s-total { background:#BBDEFB; font-weight:bold; }
    </style>
    <div style="overflow-x:auto; max-height:400px; overflow-y:auto; position:relative;">
    <table class="r1s-table">
    """

    # Header
    html += "<tr><th>קטגוריה</th>"
    for d in dates:
        day_heb = HEB_DAYS.get(d.weekday(), "")
        html += f"<th>{day_heb}<br>{d.strftime('%d/%m')}</th>"
    html += "</tr>"

    # Per category rows
    for cat in categories:
        if not any(summary_data[cat][d]["total"] > 0 for d in dates):
            continue
        html += f'<tr><td class="r1s-cat">{cat}</td>'
        for d in dates:
            pres = summary_data[cat][d]["present"]
            total = summary_data[cat][d]["total"]
            if total > 0:
                pct = round(pres / total * 100)
                bg = "#C8E6C9" if pct >= 70 else "#FFF9C4" if pct >= 40 else "#FFCDD2"
                html += f'<td style="background:{bg};">{pres}/{total}</td>'
            else:
                html += "<td>—</td>"
        html += "</tr>"

    # Total row
    html += '<tr><td class="r1s-cat r1s-total">סה"כ</td>'
    for d in dates:
        pres = total_by_date[d]["present"]
        total = total_by_date[d]["total"]
        pct_str = f" ({round(pres/total*100)}%)" if total > 0 else ""
        html += f'<td class="r1s-total">{pres}/{total}{pct_str}</td>'
    html += "</tr>"

    html += "</table></div>"
    st.markdown(html, unsafe_allow_html=True)


# ╔══════════════════════════════════════════════════════════════╗
# ║                   QUICK STATUS EDIT                          ║
# ╚══════════════════════════════════════════════════════════════╝

def _render_quick_status_edit(pid: int, p_start, p_end):
    """Quick status editing — bulk by group or per-soldier table."""
    st.markdown("### ✏️ עדכון סטטוס מהיר")

    status_options = get_status_options(pid)
    all_statuses = [s.name for s in status_options] if status_options else [
        "בבסיס", "התייצב", "חוזר מחופש", "חופש", "יוצא לחופש",
        "פיצול", "יוצא לפיצול", "גימלים", "לא בשמפ", "נפקד",
        "משתחרר", "צפוי להתייצב", "סיפוח מאוחר",
    ]
    sub_units = get_sub_units(pid)
    sub_units = [u for u in sub_units if u != IRRELEVANT_UNIT]

    tab_bulk, tab_table = st.tabs([
        "📋 עדכון קבוצתי — מחלקה / כל הפלוגה",
        "📝 עדכון פרטני — טבלה עם סטטוס לכל חייל",
    ])

    with tab_bulk:
        _render_bulk_group_update(pid, p_start, p_end, all_statuses, sub_units)
    with tab_table:
        _render_per_soldier_table(pid, p_start, p_end, all_statuses, sub_units)


def _render_bulk_group_update(pid, p_start, p_end, all_statuses, sub_units):
    """Bulk update: select whole company or specific squads, set one status for all."""
    st.markdown(
        "בחר **כל הפלוגה** או **מחלקות ספציפיות**, בחר טווח ימים וסטטוס, "
        "ועדכן את כולם בלחיצה אחת."
    )

    # ── Date range ──
    dcol1, dcol2 = st.columns(2)
    safe_default = min(max(p_start, date.today()), p_end)
    with dcol1:
        edit_start = st.date_input(
            "מתאריך", value=safe_default,
            min_value=p_start, max_value=p_end,
            key="r1_bulk_start",
        )
    with dcol2:
        edit_end = st.date_input(
            "עד תאריך", value=min(edit_start + timedelta(days=2), p_end),
            min_value=edit_start, max_value=p_end,
            key="r1_bulk_end",
        )

    # ── Group selection ──
    group_options = ["כל הפלוגה"] + sub_units
    selected_group = st.radio(
        "בחר קבוצה", group_options, horizontal=True, key="r1_bulk_group",
    )

    # If a specific squad is selected, allow multi-select for several squads
    if selected_group == "כל הפלוגה":
        selected_units = sub_units  # all
    else:
        # Pre-select the chosen squad, let user add more
        selected_units = st.multiselect(
            "מחלקות נבחרות", sub_units,
            default=[selected_group] if selected_group in sub_units else sub_units,
            key="r1_bulk_units_multi",
        )

    # ── Status ──
    CLEAR_LABEL = "🗑️ הסר סטטוס (נקה)"
    status_choices = [CLEAR_LABEL] + all_statuses
    new_status = st.selectbox("סטטוס חדש", status_choices, key="r1_bulk_status")
    is_clear = new_status == CLEAR_LABEL

    # ── Preview ──
    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)
    if selected_units:
        soldiers = [s for s in soldiers if s.get("sub_unit") in selected_units]

    num_days = (edit_end - edit_start).days + 1
    num_soldiers = len(soldiers)

    # Show breakdown by squad
    squad_counts = defaultdict(int)
    for s in soldiers:
        squad_counts[s.get("sub_unit", "ללא")] += 1

    preview_parts = [f"**{u}**: {c}" for u, c in squad_counts.items()]
    st.info(
        f"ייעודכנו **{num_soldiers}** חיילים × **{num_days}** ימים "
        f"= **{num_soldiers * num_days}** רשומות\n\n"
        + " · ".join(preview_parts)
    )

    btn_label = "🗑️ הסר סטטוס מכולם" if is_clear else "💾 עדכן סטטוס לכולם"
    btn_type = "secondary" if is_clear else "primary"
    if st.button(btn_label, type=btn_type, key="r1_bulk_save"):
        if not soldiers:
            st.error("אין חיילים בקבוצה שנבחרה")
            return
        count = 0
        current = edit_start
        while current <= edit_end:
            sids = [s["soldier_id"] for s in soldiers]
            if is_clear:
                cnt = bulk_clear_status(pid, sids, current)
            else:
                cnt = bulk_set_status(pid, sids, current, new_status)
            count += cnt
            current += timedelta(days=1)
        if is_clear:
            st.success(f"✅ הוסרו {count} סטטוסים")
        else:
            st.success(f"✅ עודכנו {count} סטטוסים")
        st.rerun()


def _render_per_soldier_table(pid, p_start, p_end, all_statuses, sub_units):
    """Per-soldier status editing — two modes: dropdown table or visual buttons."""
    st.markdown(
        "בחר תאריך וסנן לפי מחלקה — שנה סטטוס **לכל חייל בנפרד**."
    )

    # ── Date range toggle ──
    safe_default = min(max(p_start, date.today()), p_end)
    use_range = st.toggle("עריכה לטווח תאריכים", value=False, key="r1_tbl_range_toggle")

    if use_range:
        dcol1, dcol2 = st.columns(2)
        with dcol1:
            edit_start = st.date_input(
                "מתאריך", value=safe_default,
                min_value=p_start, max_value=p_end,
                key="r1_table_start",
            )
        with dcol2:
            edit_end = st.date_input(
                "עד תאריך", value=min(edit_start + timedelta(days=3), p_end),
                min_value=edit_start, max_value=p_end,
                key="r1_table_end",
            )
        date_list = []
        d = edit_start
        while d <= edit_end:
            date_list.append(d)
            d += timedelta(days=1)
    else:
        edit_start = st.date_input(
            "תאריך", value=safe_default,
            min_value=p_start, max_value=p_end,
            key="r1_table_date",
        )
        edit_end = edit_start
        date_list = [edit_start]

    num_days = len(date_list)
    if num_days > 14:
        st.warning("מומלץ לעבוד עם עד 14 ימים בכל פעם")

    # ── Sub-unit filter ──
    if sub_units:
        selected_units = st.multiselect(
            "סנן לפי מחלקה (השאר ריק לכולם)", sub_units,
            default=sub_units, key="r1_table_units",
        )
    else:
        selected_units = []

    # ── Load soldiers + current statuses ──
    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)
    if selected_units:
        soldiers = [s for s in soldiers if s.get("sub_unit") in selected_units]

    if not soldiers:
        st.info("אין חיילים להצגה")
        return

    grid = get_daily_status_grid(pid, edit_start, edit_end)
    full_status_map: dict[tuple[int, str], str] = {}
    full_notes_map: dict[tuple[int, str], str] = {}
    for sol in grid["soldiers"]:
        for dk, sv in sol["statuses"].items():
            full_status_map[(sol["soldier_id"], dk)] = sv
        for dk, nv in sol.get("notes", {}).items():
            full_notes_map[(sol["soldier_id"], dk)] = nv

    # ── Two editing modes ──
    mode_table, mode_visual = st.tabs([
        "📝 טבלת בחירה",
        "🎨 כפתורים ויזואליים",
    ])

    with mode_table:
        _render_dropdown_table(
            pid, soldiers, all_statuses, date_list, full_status_map, num_days,
            full_notes_map,
        )

    with mode_visual:
        _render_visual_buttons(
            pid, soldiers, all_statuses, date_list, full_status_map,
            full_notes_map,
        )


def _render_dropdown_table(
    pid, soldiers, all_statuses, date_list, full_status_map, num_days,
    full_notes_map=None,
):
    """Dropdown-based data editor table — one table per unit with independent save."""
    # Status color mapping for the colored preview
    _STATUS_COLORS = {
        "בבסיס": ("#E8F5E9", "#1B5E20"),
        "התייצב": ("#C8E6C9", "#1B5E20"),
        "חוזר מחופש": ("#DCEDC8", "#33691E"),
        "צפוי להתייצב": ("#F0F4C3", "#33691E"),
        "סיפוח מאוחר": ("#FFF9C4", "#F57F17"),
        "חופש": ("#FFCDD2", "#B71C1C"),
        "יוצא לחופש": ("#FFAB91", "#BF360C"),
        "פיצול": ("#FFE0B2", "#E65100"),
        "יוצא לפיצול": ("#FFE0B2", "#E65100"),
        "גימלים": ("#F8BBD0", "#880E4F"),
        "משתחרר": ("#E1BEE7", "#4A148C"),
        "לא בשמפ": ("#CFD8DC", "#37474F"),
        "נפקד": ("#FF8A80", "#B71C1C"),
        "יוצא לקורס": ("#D7CCC8", "#4E342E"),
        "רספ/סרספ": ("#C5CAE9", "#1A237E"),
        "סמבצים": ("#B2DFDB", "#004D40"),
        "סוואנה": ("#B2EBF2", "#006064"),
    }

    # Options include empty so unsaved soldiers are visually clear
    status_options_with_empty = [""] + all_statuses

    # Group soldiers by sub_unit
    units: dict[str, list[dict]] = defaultdict(list)
    for s in soldiers:
        units[s.get("sub_unit", "ללא מחלקה")].append(s)

    # Build date column labels once
    date_col_labels = []
    for d in date_list:
        heb_day = HEB_DAYS.get(d.weekday(), "")
        col_label = f"{heb_day} {d.day}/{d.month:02d}"
        date_col_labels.append((d, col_label))

    # Build note column labels matching date columns
    note_col_labels = []
    for d, col_label in date_col_labels:
        note_col_labels.append((d, f"📝 {col_label}"))

    for unit_name in sorted(units.keys()):
        unit_soldiers = units[unit_name]
        with st.expander(f"📁 {unit_name} ({len(unit_soldiers)} חיילים)", expanded=True):
            # Build DataFrame for this unit
            rows = []
            for s in unit_soldiers:
                row = {
                    "soldier_id": s["soldier_id"],
                    "שם": s["full_name"],
                    "תפקיד": s.get("role", ""),
                }
                for d, col_label in date_col_labels:
                    dk = d.isoformat()
                    cur = full_status_map.get((s["soldier_id"], dk), "")
                    row[col_label] = cur if cur in all_statuses else ""
                for d, note_label in note_col_labels:
                    dk = d.isoformat()
                    row[note_label] = (full_notes_map or {}).get((s["soldier_id"], dk), "")
                rows.append(row)

            df = pd.DataFrame(rows)

            date_cols = [cl for _, cl in date_col_labels]

            # ── Editable table ──
            safe_key = unit_name.replace(" ", "_").replace('"', '')

            # Build column config
            col_config = {
                "soldier_id": None,
                "שם": st.column_config.TextColumn("שם", disabled=True, width="medium", pinned=True),
                "תפקיד": st.column_config.TextColumn("תפקיד", disabled=True, width="small", pinned=True),
            }
            for _, col_label in date_col_labels:
                col_config[col_label] = st.column_config.SelectboxColumn(
                    col_label,
                    options=status_options_with_empty,
                    required=False,
                    width="small" if num_days > 3 else "medium",
                )
            for _, note_label in note_col_labels:
                col_config[note_label] = st.column_config.TextColumn(
                    note_label,
                    width="small",
                )

            table_height = min(35 * len(unit_soldiers) + 60, 600)
            table_height = max(table_height, 200)

            edited = st.data_editor(
                df,
                column_config=col_config,
                hide_index=True,
                use_container_width=True,
                height=table_height,
                key=f"r1_tbl_{safe_key}",
            )

            # ── Colored read-only preview (synced with edits) ──
            with st.expander("🎨 תצוגה צבעונית", expanded=False):
                def _style_status(val):
                    if not val or val not in _STATUS_COLORS:
                        return ""
                    bg, fg = _STATUS_COLORS[val]
                    return f"background-color: {bg}; color: {fg}; font-weight: bold; text-align: center;"

                display_df = edited[["שם", "תפקיד"] + date_cols].copy()
                styled = display_df.style.map(_style_status, subset=date_cols)
                st.dataframe(
                    styled,
                    hide_index=True,
                    use_container_width=True,
                    height=min(35 * len(unit_soldiers) + 60, 500),
                )

            # ── Leave policy check ──
            _check_leave_warnings(pid, edited, date_col_labels, all_statuses, unit_key=safe_key)

            # ── Per-unit save button ──
            if st.button(f"💾 שמור {unit_name}", type="primary", key=f"r1_save_{safe_key}"):
                count = 0
                for _, row in edited.iterrows():
                    sid = int(row["soldier_id"])
                    for d, col_label in date_col_labels:
                        new_st = row[col_label]
                        dk = d.isoformat()
                        cur_st = full_status_map.get((sid, dk), "")
                        if new_st and new_st != cur_st:
                            set_status(pid, sid, d, new_st)
                            count += 1
                        elif not new_st and cur_st:
                            from military_manager.services.status_service import bulk_clear_status
                            bulk_clear_status(pid, [sid], d)
                            count += 1
                    # Save notes
                    for d, note_label in note_col_labels:
                        dk = d.isoformat()
                        new_note = row.get(note_label, "") or ""
                        cur_note = (full_notes_map or {}).get((sid, dk), "")
                        if new_note != cur_note:
                            set_status_notes(pid, sid, d, new_note)
                            count += 1
                if count:
                    st.success(f"✅ {unit_name}: עודכנו {count} שינויים")
                    st.rerun()
                else:
                    st.info(f"אין שינויים ב{unit_name}")


# ── Status button colors ──
STATUS_BTN_COLORS = {
    "בבסיס": ("#2E7D32", "#C8E6C9", "#1B5E20"),       # dark green bg, light green, text
    "התייצב": ("#388E3C", "#A5D6A7", "#1B5E20"),
    "חוזר מחופש": ("#43A047", "#E8F5E9", "#2E7D32"),
    "צפוי להתייצב": ("#558B2F", "#DCEDC8", "#33691E"),
    "סיפוח מאוחר": ("#827717", "#F0F4C3", "#33691E"),
    "חופש": ("#C62828", "#FFCDD2", "#B71C1C"),
    "יוצא לחופש": ("#D84315", "#FFAB91", "#BF360C"),
    "פיצול": ("#E65100", "#FFE0B2", "#BF360C"),
    "יוצא לפיצול": ("#EF6C00", "#FFE0B2", "#E65100"),
    "גימלים": ("#AD1457", "#F8BBD0", "#880E4F"),
    "משתחרר": ("#6A1B9A", "#E1BEE7", "#4A148C"),
    "לא בשמפ": ("#546E7A", "#CFD8DC", "#37474F"),
    "נפקד": ("#D32F2F", "#FF8A80", "#B71C1C"),
}


def _render_visual_buttons(
    pid, soldiers, all_statuses, date_list, full_status_map,
    full_notes_map=None,
):
    """Visual colored-pill status editing using st.pills — horizontal layout."""

    st.caption("בחר סטטוס לכל חייל. השינויים נשמרים בלחיצה על 'שמור'.")

    # ── Inject CSS to color each pill by its position (= status index) ──
    css_parts = ["<style>"]
    # Make pills RTL and compact
    css_parts.append("""
    [data-testid="stPills"] > div > div {
        direction: rtl;
        flex-wrap: wrap;
        gap: 3px !important;
    }
    [data-testid="stPills"] > div > div > button {
        font-size: 12px !important;
        padding: 3px 10px !important;
        min-height: unset !important;
        line-height: 1.3 !important;
    }
    [data-testid="stPills"] label { direction: rtl; text-align: right; }
    """)

    for i, status in enumerate(all_statuses):
        active_bg, light_bg, text_color = STATUS_BTN_COLORS.get(
            status, ("#555", "#eee", "#333")
        )
        n = i + 1  # nth-child is 1-indexed
        # Unselected pill — light background with tinted border
        css_parts.append(f"""
        [data-testid="stPills"] > div > div > button:nth-child({n}) {{
            background-color: {light_bg} !important;
            color: {text_color} !important;
            border: 1px solid {active_bg}55 !important;
        }}
        """)
        # Selected pill — bold colored background
        css_parts.append(f"""
        [data-testid="stPills"] > div > div > button:nth-child({n})[aria-checked="true"],
        [data-testid="stPills"] > div > div > button:nth-child({n})[data-active="true"] {{
            background-color: {active_bg} !important;
            color: white !important;
            border-color: {text_color} !important;
            font-weight: bold !important;
            box-shadow: 0 0 0 2px {active_bg}44 !important;
        }}
        """)

    css_parts.append("</style>")
    st.markdown("\n".join(css_parts), unsafe_allow_html=True)

    st.markdown("---")

    # Group soldiers by sub_unit for organized display
    units: dict[str, list[dict]] = defaultdict(list)
    for s in soldiers:
        units[s.get("sub_unit", "ללא מחלקה")].append(s)

    for unit_name in sorted(units.keys()):
        unit_soldiers = units[unit_name]
        # Collect changes per unit
        unit_changes: list[dict] = []

        with st.expander(f"📁 {unit_name} ({len(unit_soldiers)} חיילים)", expanded=True):
            for s in unit_soldiers:
                sid = s["soldier_id"]
                name = s.get("full_name", "?")
                role = s.get("role", "") or ""

                for d in date_list:
                    dk = d.isoformat()
                    cur_status = full_status_map.get((sid, dk), "")
                    cur_note = (full_notes_map or {}).get((sid, dk), "")

                    if len(date_list) > 1:
                        heb_day = HEB_DAYS.get(d.weekday(), "")
                        label = f"{name} · {role} — {heb_day} {d.day}/{d.month:02d}"
                    else:
                        label = f"{name} · {role}"

                    # Default to None (no selection) if no status set
                    default_val = (
                        cur_status if cur_status in all_statuses
                        else None
                    )

                    # Status pills + note icon in one row
                    pills_col, note_col = st.columns([12, 1])

                    with pills_col:
                        new_status = st.pills(
                            label,
                            all_statuses,
                            default=default_val,
                            key=f"vp_{sid}_{dk}",
                        )

                    with note_col:
                        note_icon = "📝" if cur_note else "💬"
                        with st.popover(note_icon, help="הערה"):
                            st.markdown(f"**הערה עבור {name}**")
                            new_note = st.text_input(
                                "הערה",
                                value=cur_note,
                                key=f"note_{sid}_{dk}",
                                placeholder="סיבת היעדרות, פרטים...",
                                label_visibility="collapsed",
                            )
                            if new_note != cur_note:
                                if st.button("💾", key=f"note_save_{sid}_{dk}"):
                                    set_status_notes(pid, sid, d, new_note)
                                    st.rerun()

                    # Track changes
                    if new_status and new_status != cur_status:
                        unit_changes.append({
                            "sid": sid, "date": d, "status": new_status,
                            "name": name, "old": cur_status,
                        })
                    elif not new_status and cur_status:
                        unit_changes.append({
                            "sid": sid, "date": d, "status": "",
                            "name": name, "old": cur_status,
                        })

                # Thin separator between soldiers
                st.markdown(
                    "<hr style='margin:1px 0; border:none; border-top:1px solid #eee;'>",
                    unsafe_allow_html=True,
                )

            # ── Prominent save button — sticky at bottom of each unit ──
            if unit_changes:
                st.markdown(
                    f'<div style="background:#FFF3E0; border:2px solid #FF9800; '
                    f'border-radius:10px; padding:10px; text-align:center; '
                    f'margin:8px 0;">'
                    f'<b style="color:#E65100; font-size:1.1em;">'
                    f'⚠️ {len(unit_changes)} שינויים לא שמורים ב{unit_name}</b>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                for ch in unit_changes:
                    old_bg = STATUS_BTN_COLORS.get(ch["old"], ("#555", "#eee", "#333"))[0]
                    new_bg = STATUS_BTN_COLORS.get(ch["status"], ("#555", "#eee", "#333"))[0]
                    st.markdown(
                        f"- **{ch['name']}** ({ch['date'].strftime('%d/%m')}): "
                        f"<span style='background:{old_bg}; color:white; padding:1px 6px; "
                        f"border-radius:6px; font-size:11px;'>{ch['old'] or '—'}</span>"
                        f" → "
                        f"<span style='background:{new_bg}; color:white; padding:1px 6px; "
                        f"border-radius:6px; font-size:11px;'>{ch['status'] or '🗑️ מחיקה'}</span>",
                        unsafe_allow_html=True,
                    )

                safe_key = unit_name.replace(' ', '_').replace('"', '')
                if st.button(
                    f"💾 שמור {len(unit_changes)} שינויים — {unit_name}",
                    type="primary",
                    key=f"r1_vis_save_{safe_key}",
                    use_container_width=True,
                ):
                    count = 0
                    for ch in unit_changes:
                        if ch["status"]:
                            set_status(pid, ch["sid"], ch["date"], ch["status"])
                        else:
                            bulk_clear_status(pid, [ch["sid"]], ch["date"])
                        count += 1
                    st.success(f"✅ {unit_name}: עודכנו {count} סטטוסים")
                    st.rerun()


# ╔══════════════════════════════════════════════════════════════╗
# ║              LEAVE POLICY — 4/10 RULE                        ║
# ╚══════════════════════════════════════════════════════════════╝

# 4 home days per 10 base days
LEAVE_MAX_DAYS = 4
LEAVE_CYCLE_DAYS = 10
HOME_STATUSES = {"חופש", "פיצול", "גימלים"}  # "full" home days


def _check_leave_warnings(pid: int, edited_df, date_col_labels: list, all_statuses: list, unit_key: str = ""):
    """Check if any soldier exceeds the 4/10 leave policy.

    For each soldier, count how many 'home' days they have in the edited table
    *plus* existing history in the period. Warn if > 4 per 10.
    """
    if edited_df is None or edited_df.empty or not date_col_labels:
        return

    # Get full status history for each soldier in the edited table
    violations = []

    for _, row in edited_df.iterrows():
        sid = int(row["soldier_id"])
        name = row["שם"]

        # Count home days in the edited range
        new_home_days = 0
        for _, col_label in date_col_labels:
            val = row.get(col_label, "")
            if val in HOME_STATUSES:
                new_home_days += 1

        # Get existing history to calculate total
        history = get_soldier_status_history(pid, sid)
        edited_dates = {d.isoformat() for d, _ in date_col_labels}

        # Total home days: existing (excluding edited dates) + new edited ones
        existing_home = sum(
            1 for h in history
            if h["status"] in HOME_STATUSES and h["date"].isoformat() not in edited_dates
        )
        existing_base = sum(
            1 for h in history
            if h["status"] in PRESENT_STATUSES and h["date"].isoformat() not in edited_dates
        )

        # Count base days in new edits too
        new_base_days = 0
        for _, col_label in date_col_labels:
            val = row.get(col_label, "")
            if val in PRESENT_STATUSES:
                new_base_days += 1

        total_home = existing_home + new_home_days
        total_base = existing_base + new_base_days
        total_period = total_base + total_home

        # Calculate allowed home days: 4 per 10 base days (proportional)
        if total_period > 0:
            allowed = max(LEAVE_MAX_DAYS, (total_base // LEAVE_CYCLE_DAYS) * LEAVE_MAX_DAYS + LEAVE_MAX_DAYS)
        else:
            allowed = LEAVE_MAX_DAYS

        # Simple check: if total home > 4, warn
        if total_home > LEAVE_MAX_DAYS:
            excess = total_home - LEAVE_MAX_DAYS
            violations.append({
                "soldier_id": sid,
                "name": name,
                "total_home": total_home,
                "total_base": total_base,
                "excess": excess,
            })

    if not violations:
        return

    # ── Show warnings ──
    st.markdown("---")
    st.markdown("### ⚠️ חריגה ממדיניות חופשות (4 ימי בית ל-10 ימי בסיס)")
    for v in violations:
        st.warning(
            f"**{v['name']}** — {v['total_home']} ימי בית "
            f"(חריגה של {v['excess']} ימים). "
            f"נדרש אישור מ\"פ."
        )

    # ── Send approval request ──
    st.markdown("**📨 שלח בקשת אישור למ\"פ:**")
    soldiers_list = [v["name"] for v in violations]
    st.caption(f"חיילים חורגים: {', '.join(soldiers_list)}")
    reason = st.text_area(
        "סיבה / הערות לבקשה",
        placeholder="למשל: חייל עם אירוע משפחתי מיוחד...",
        key=f"r1_leave_request_reason_{unit_key}",
    )

    if st.button("📨 שלח בקשת אישור למ\"פ", key=f"r1_leave_request_send_{unit_key}"):
        from sqlalchemy import select as sa_select
        count = 0
        for v in violations:
            # Check if there's already a pending request for this soldier
            with get_session() as session:
                existing = session.execute(
                    sa_select(Request).where(
                        Request.period_id == pid,
                        Request.soldier_id == v["soldier_id"],
                        Request.request_type == "חריגת חופשה",
                        Request.status == "ממתין",
                    )
                ).scalar_one_or_none()
                if existing:
                    continue  # skip duplicate

                req = Request(
                    period_id=pid,
                    soldier_id=v["soldier_id"],
                    request_type="חריגת חופשה",
                    subject=f"חריגה: {v['total_home']} ימי בית (מותר {LEAVE_MAX_DAYS})",
                    details=reason or "",
                    status="ממתין",
                    reason=f"סה\"כ {v['total_home']} ימי בית, "
                           f"{v['total_base']} ימי בסיס. חריגה: {v['excess']} ימים.",
                )
                session.add(req)
                session.commit()
                count += 1
        if count:
            st.success(f"✅ נשלחו {count} בקשות אישור למ\"פ — ניתן לאשר בעמוד 'בקשות'")
        else:
            st.info("כבר קיימות בקשות ממתינות לחיילים אלו")


# ╔══════════════════════════════════════════════════════════════╗
# ║                   MANPOWER SUMMARY                           ║
# ╚══════════════════════════════════════════════════════════════╝

def _render_manpower_summary(pid: int, p_start, p_end):
    """Manpower summary — per day: how many of each role category are present."""
    st.markdown("### 📈 סיכום כוח אדם — גרף")

    import plotly.express as px

    col1, col2 = st.columns(2)
    with col1:
        chart_start = st.date_input(
            "מתאריך",
            value=min(max(p_start, date.today()), p_end),
            min_value=p_start,
            max_value=p_end,
            key="r1_chart_start",
        )
    with col2:
        chart_end = st.date_input(
            "עד תאריך",
            value=min(chart_start + timedelta(days=13), p_end),
            min_value=chart_start,
            max_value=p_end,
            key="r1_chart_end",
        )

    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)
    grid = get_daily_status_grid(pid, chart_start, chart_end)
    status_map = grid.get("statuses", {}) if grid else {}

    dates = []
    current = chart_start
    while current <= chart_end:
        dates.append(current)
        current += timedelta(days=1)

    # Build chart data
    chart_rows = []
    for d in dates:
        present_count = 0
        away_count = 0
        absent_count = 0
        no_status = 0

        for s in soldiers:
            key = f"{s['soldier_id']}_{d.isoformat()}"
            status = status_map.get(key, "")
            if status in PRESENT_STATUSES:
                present_count += 1
            elif status in AWAY_STATUSES:
                away_count += 1
            elif status in ABSENT_STATUSES:
                absent_count += 1
            else:
                no_status += 1

        day_heb = HEB_DAYS.get(d.weekday(), "")
        date_label = f"{day_heb} {d.strftime('%d/%m')}"
        chart_rows.append({"תאריך": date_label, "קטגוריה": "נוכחים", "כמות": present_count})
        chart_rows.append({"תאריך": date_label, "קטגוריה": "חופשות/פיצולים", "כמות": away_count})
        chart_rows.append({"תאריך": date_label, "קטגוריה": "לא בשמ\"פ/נפקד", "כמות": absent_count})
        if no_status > 0:
            chart_rows.append({"תאריך": date_label, "קטגוריה": "ללא סטטוס", "כמות": no_status})

    if not chart_rows:
        st.info("אין נתונים")
        return

    df = pd.DataFrame(chart_rows)

    fig = px.bar(
        df,
        x="תאריך",
        y="כמות",
        color="קטגוריה",
        barmode="stack",
        color_discrete_map={
            "נוכחים": "#4CAF50",
            "חופשות/פיצולים": "#FF9800",
            "לא בשמ\"פ/נפקד": "#F44336",
            "ללא סטטוס": "#9E9E9E",
        },
        title="כוח אדם נוכח לפי ימים",
    )
    fig.update_layout(
        font=dict(family="Segoe UI, Arial"),
        xaxis_title="",
        yaxis_title="מספר חיילים",
        height=400,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    st.plotly_chart(fig, use_container_width=True)

    # Also show the numbers table
    st.markdown("#### 📋 טבלת נוכחות")

    summary_rows = []
    for d in dates:
        present = sum(
            1 for s in soldiers
            if status_map.get(f"{s['soldier_id']}_{d.isoformat()}", "") in PRESENT_STATUSES
        )
        away = sum(
            1 for s in soldiers
            if status_map.get(f"{s['soldier_id']}_{d.isoformat()}", "") in AWAY_STATUSES
        )
        absent = sum(
            1 for s in soldiers
            if status_map.get(f"{s['soldier_id']}_{d.isoformat()}", "") in ABSENT_STATUSES
        )
        total = len(soldiers)
        pct = round(present / total * 100) if total > 0 else 0

        day_heb = HEB_DAYS.get(d.weekday(), "")
        summary_rows.append({
            "תאריך": f"{day_heb} {d.strftime('%d/%m/%Y')}",
            "נוכחים": present,
            "חופשות": away,
            "לא בשמ\"פ": absent,
            "ללא סטטוס": total - present - away - absent,
            "סה\"כ": total,
            "% נוכחות": f"{pct}%",
        })

    df_summary = pd.DataFrame(summary_rows)
    st.dataframe(df_summary, use_container_width=True, hide_index=True)


def _render_group_percentage_report(pid: int, p_start, p_end):
    """Dynamic group percentage report — shows daily group stats for a date range."""
    from military_manager.services.stats_service import (
        compute_percentages, init_default_groups, get_status_groups, get_setting,
    )

    st.markdown("### 📊 אחוזים לפי קבוצות סטטוס")
    st.caption(
        "חישוב דינמי: כמות חיילים בכל קבוצה, אחוזים מתוך סה\"כ בשמ\"פ. "
        "ניתן להגדיר קבוצות בהגדרות ← קבוצות סטטוס."
    )

    init_default_groups(pid)
    groups = get_status_groups(pid)
    group_names = [g["name"] for g in groups]
    threshold = float(get_setting(pid, "home_alert_percent", "25"))

    # Date range
    col1, col2 = st.columns(2)
    with col1:
        grp_start = st.date_input(
            "מתאריך",
            value=min(max(p_start, date.today()), p_end),
            min_value=p_start,
            max_value=p_end,
            key="r1_grp_start",
        )
    with col2:
        grp_end = st.date_input(
            "עד תאריך",
            value=min(grp_start + timedelta(days=6), p_end),
            min_value=grp_start,
            max_value=p_end,
            key="r1_grp_end",
        )

    # Today's snapshot KPIs
    today = date.today()
    if p_start <= today <= p_end:
        day_stats = compute_percentages(pid, today)
        if day_stats and day_stats["groups"]:
            st.markdown("#### 📊 מצב נוכחי — היום")
            grp_cols = st.columns(min(len(day_stats["groups"]), 5))
            for i, (gname, gdata) in enumerate(day_stats["groups"].items()):
                pct = gdata["percent"]
                cnt = gdata["count"]
                color = gdata.get("color", "#9E9E9E")
                is_alert = gname == "בחופש" and pct > threshold
                with grp_cols[i % len(grp_cols)]:
                    if is_alert:
                        st.markdown(
                            f'<div style="background:#FFCDD2;padding:10px;border-radius:8px;text-align:center;'
                            f'border:2px solid #B71C1C;">'
                            f'<b style="color:#B71C1C;font-size:1.4em;">{pct}%</b><br>'
                            f'<span style="color:#B71C1C;">{gname}: {cnt}</span><br>'
                            f'<span style="color:#B71C1C;font-size:0.75em;">⚠️ מעל {threshold}%</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            f'<div style="background:{color}22;padding:10px;border-radius:8px;text-align:center;'
                            f'border:1px solid {color};">'
                            f'<b style="color:{color};font-size:1.4em;">{pct}%</b><br>'
                            f'<span>{gname}: {cnt}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
            for alert in day_stats.get("alerts", []):
                st.error(alert["message"])

    # Daily table
    st.markdown("---")
    st.markdown("#### 📋 טבלת אחוזים יומית")

    dates = []
    current = grp_start
    while current <= grp_end:
        dates.append(current)
        current += timedelta(days=1)

    rows = []
    for d in dates:
        day_heb = HEB_DAYS.get(d.weekday(), "")
        row = {"תאריך": f"{day_heb} {d.strftime('%d/%m')}"}
        stats = compute_percentages(pid, d)
        for gname in group_names:
            gdata = stats["groups"].get(gname, {"count": 0, "percent": 0})
            row[f"{gname} (כמות)"] = gdata["count"]
            row[f"{gname} (%)"] = gdata["percent"]
        row["סה\"כ בשמ\"פ"] = stats["total_in_shmap"]
        rows.append(row)

    if rows:
        df = pd.DataFrame(rows)

        # Style leave % columns in red when over threshold
        def _red_if_over(val):
            if isinstance(val, (int, float)) and val > threshold:
                return "background-color: #FFCDD2; color: #B71C1C; font-weight: bold"
            return ""

        leave_pct_col = "בחופש (%)"
        if leave_pct_col in df.columns:
            styled = df.style.map(_red_if_over, subset=[leave_pct_col])
        else:
            styled = df.style

        st.dataframe(styled, use_container_width=True, hide_index=True)
    else:
        st.info("אין נתונים")


# ╔══════════════════════════════════════════════════════════════╗
# ║                   STUDENT REPORT                             ║
# ╚══════════════════════════════════════════════════════════════╝

# Statuses that mean the soldier has NOT yet enlisted (don't count for service)
_NOT_ENLISTED_STATUSES = {"לא בשמפ", "לא בשמ\"פ", "פיצול", "סיפוח מאוחר"}

# "פיצול" is also tracked separately because it extends the release date
_SPLIT_STATUSES = {"פיצול", "יוצא לפיצול"}

STUDENT_DISCOUNT = 0.25  # 25% reduction


def _compute_student_service(
    soldier: dict,
    p_start: date,
    p_end: date,
    status_map: dict,
) -> dict:
    """Compute service stats for a student soldier.

    Logic:
    - Total order days = p_end - p_start + 1
    - Required service days = ceil(75% of order days)
    - Count *actual* service days from Report 1:
      * Before first enlistment day: skip "לא בשמפ", "פיצול", "סיפוח מאוחר"
      * Once soldier appears with ANY other status → enlisted
      * After enlistment: every day counts (including חופש)
    - Count פיצול days (total, including before & after enlistment)
    - Expected release date = p_start + required_days + פיצול days
      (capped at p_end)
    """
    import math

    sid = soldier["soldier_id"]
    total_order_days = (p_end - p_start).days + 1
    required_days = math.ceil(total_order_days * (1 - STUDENT_DISCOUNT))

    # Walk through all dates in period
    enlisted = False
    enlistment_date = None
    service_days = 0
    split_days = 0

    d = p_start
    while d <= p_end:
        key = f"{sid}_{d.isoformat()}"
        status = status_map.get(key, "")

        # Track split days always
        if status in _SPLIT_STATUSES:
            split_days += 1

        if not enlisted:
            # Check if this day = enlistment (status exists and is NOT one of the
            # "not yet enlisted" statuses)
            if status and status not in _NOT_ENLISTED_STATUSES:
                enlisted = True
                enlistment_date = d
                service_days += 1
            # else: not enlisted yet, day doesn't count
        else:
            # After enlistment, every day counts
            service_days += 1

        d += timedelta(days=1)

    # Expected release date
    # The soldier needs `required_days` of actual service.
    # Split days extend the total calendar span needed.
    if enlistment_date:
        # Walk forward from enlistment_date counting service days
        # until we reach required_days
        counted = 0
        rd = enlistment_date
        release_date = p_end  # fallback
        while rd <= p_end:
            rkey = f"{sid}_{rd.isoformat()}"
            rstatus = status_map.get(rkey, "")

            if rstatus in _SPLIT_STATUSES:
                # Split day — doesn't count as service, extends window
                pass
            else:
                counted += 1

            if counted >= required_days:
                release_date = rd
                break
            rd += timedelta(days=1)
        else:
            release_date = p_end

        # Cap at period end
        if release_date > p_end:
            release_date = p_end
    else:
        release_date = None

    remaining = max(0, required_days - service_days + split_days) if enlisted else required_days

    return {
        "soldier": soldier,
        "total_order_days": total_order_days,
        "required_days": required_days,
        "service_days": service_days,
        "split_days": split_days,
        "enlisted": enlisted,
        "enlistment_date": enlistment_date,
        "release_date": release_date,
        "remaining": remaining,
    }


def _render_student_report(pid: int, p_start, p_end):
    """Student service report — shows service days & expected release per student."""
    st.markdown("### 🎓 דו\"ח סטודנטים — קיצור שירות")
    st.caption(
        "סטודנטים זכאים לקצר שירות מילואים ב-25% מאורך הצו. "
        "הדו\"ח מחשב ימי שירות בפועל ותאריך שחרור צפוי."
    )

    soldiers = get_period_soldiers(pid, exclude_irrelevant_unit=True)

    # Filter students only
    students = [s for s in soldiers if s.get("is_student")]
    if not students:
        st.info(
            "אין חיילים המסומנים כסטודנטים. "
            "ניתן לסמן חייל כסטודנט דרך עמוד 'חיילים' ← עריכת חייל."
        )
        return

    # Load full status grid for the entire period
    grid = get_daily_status_grid(pid, p_start, p_end)
    status_map = grid.get("statuses", {}) if grid else {}

    # Compute stats
    stats = []
    for s in students:
        stat = _compute_student_service(s, p_start, p_end, status_map)
        stats.append(stat)

    # Group by sub_unit
    units: dict[str, list[dict]] = defaultdict(list)
    for stat in stats:
        u = stat["soldier"]["sub_unit"] or "ללא מחלקה"
        units[u].append(stat)

    # ── Summary KPIs ──
    total_students = len(students)
    active_short = sum(1 for s in students if s.get("student_short_service"))
    st.markdown(
        f"**סה\"כ סטודנטים:** {total_students} &nbsp;•&nbsp; "
        f"**מקצרים שירות:** {active_short} &nbsp;•&nbsp; "
        f"**אורך הצו:** {(p_end - p_start).days + 1} ימים &nbsp;•&nbsp; "
        f"**שירות נדרש (75%):** {stats[0]['required_days']} ימים"
    )
    st.markdown("---")

    # ── Per-unit tables ──
    for unit_name in sorted(units.keys()):
        unit_stats = units[unit_name]
        with st.expander(f"📁 {unit_name} ({len(unit_stats)} סטודנטים)", expanded=True):
            # Build HTML table
            html = _build_student_html_table(unit_stats, p_end)
            st.markdown(html, unsafe_allow_html=True)

    # ── Legend ──
    st.markdown("---")
    st.markdown("**מקרא:**")
    cols = st.columns(4)
    cols[0].markdown(
        '<span style="background:#C8E6C9;padding:3px 8px;border-radius:4px;">שחרור צפוי עד 3 ימים</span>',
        unsafe_allow_html=True,
    )
    cols[1].markdown(
        '<span style="background:#FFF9C4;padding:3px 8px;border-radius:4px;">שחרור צפוי עד 7 ימים</span>',
        unsafe_allow_html=True,
    )
    cols[2].markdown(
        '<span style="background:#FFCDD2;padding:3px 8px;border-radius:4px;">שחרור בעוד 7+ ימים</span>',
        unsafe_allow_html=True,
    )
    cols[3].markdown(
        '<span style="background:#E1BEE7;padding:3px 8px;border-radius:4px;">לא מקצר שירות</span>',
        unsafe_allow_html=True,
    )


def _build_student_html_table(unit_stats: list[dict], p_end: date) -> str:
    """Build HTML table for student service data."""
    today = date.today()
    html = """
    <style>
    .stu-table { width:100%; border-collapse:separate; border-spacing:0;
        direction:rtl; font-size:12px; }
    .stu-table th { background:#1565C0; color:white; padding:6px 8px;
        text-align:center; border:1px solid #aaa; font-size:11px; }
    .stu-table td { padding:5px 8px; border:1px solid #ddd; text-align:center;
        font-size:11px; }
    .stu-name { text-align:right !important; font-weight:500; min-width:110px; }
    .stu-good { background:#C8E6C9; }
    .stu-warn { background:#FFF9C4; }
    .stu-bad { background:#FFCDD2; }
    .stu-inactive { background:#E1BEE7; }
    </style>
    <table class="stu-table">
    <tr>
        <th>שם</th>
        <th>תפקיד</th>
        <th>מקצר שירות</th>
        <th>תאריך גיוס</th>
        <th>ימי שירות</th>
        <th>ימי פיצול</th>
        <th>נדרש (75%)</th>
        <th>נותרו</th>
        <th>תאריך שחרור צפוי</th>
    </tr>
    """

    for stat in unit_stats:
        s = stat["soldier"]
        name = s.get("full_name", "?")
        role = s.get("role", "") or ""
        wants_short = s.get("student_short_service", False)

        enlist_str = stat["enlistment_date"].strftime("%d/%m") if stat["enlistment_date"] else "טרם התגייס"
        service_str = str(stat["service_days"])
        split_str = str(stat["split_days"])
        required_str = str(stat["required_days"])

        if wants_short and stat["release_date"]:
            release_str = stat["release_date"].strftime("%d/%m/%Y")
            days_left = (stat["release_date"] - today).days
            if days_left <= 0:
                cls = "stu-good"
                remaining_str = "הושלם ✅"
            elif days_left <= 3:
                cls = "stu-good"
                remaining_str = f"{days_left} ימים"
            elif days_left <= 7:
                cls = "stu-warn"
                remaining_str = f"{days_left} ימים"
            else:
                cls = "stu-bad"
                remaining_str = f"{days_left} ימים"
        elif wants_short:
            release_str = "—"
            remaining_str = f"{stat['remaining']} ימים"
            cls = "stu-bad"
        else:
            # Student but not using short service
            release_str = p_end.strftime("%d/%m/%Y")
            remaining_str = "לא מקצר"
            cls = "stu-inactive"

        short_label = "✅ כן" if wants_short else "❌ לא"

        html += f"""<tr>
            <td class="stu-name">{name}</td>
            <td>{role}</td>
            <td>{short_label}</td>
            <td>{enlist_str}</td>
            <td>{service_str}</td>
            <td>{split_str}</td>
            <td>{required_str}</td>
            <td class="{cls}">{remaining_str}</td>
            <td class="{cls}"><b>{release_str}</b></td>
        </tr>"""

    html += "</table>"
    return html
